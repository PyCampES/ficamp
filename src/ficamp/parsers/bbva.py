from pathlib import Path
import os
from ficamp.types import Tx
from ficamp.parsers.protocol import Parser
from openpyxl import load_workbook


class BBVAParser(Parser):
    """Parser for BBVA bank"""

    def load(self, filename: Path | None = None):
        # filename = Path("../data/enero-febrero-bbva-cuenta.xlsx")

        filename = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data/enero-febrero-bbva-cuenta.xlsx")

        wb = load_workbook(filename)
        sheet = wb.active
        start_row = 6
        start_column = 2

        for row in sheet.iter_rows(min_row=start_row, min_col=start_column, values_only=True):
            print(row)


    def parse(self) -> list[Tx]: ...


if __name__ == "__main__":
    bbva = BBVAParser()
    bbva.load()